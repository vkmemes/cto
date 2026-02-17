#!/usr/bin/env python3
"""
Excel to schedule.json converter
Converts Excel schedule from ЯГК format to JSON format for ЯГК Schedule system.

Supports loading from GitHub URL or local file.
Implements FSM logic for detecting even/odd week patterns.
"""

import json
import re
from io import BytesIO
from typing import Dict, List, Optional, Any, Tuple

try:
    import openpyxl
    from openpyxl.cell.cell import Cell
except ImportError:
    raise ImportError("openpyxl not installed. Run: pip install openpyxl")


def clean_text(text: Any) -> str:
    """Clean and normalize text."""
    if text is None:
        return ""
    return str(text).strip()


def is_group_header(cell_text: str) -> bool:
    """
    Check if cell contains a group header pattern.
    Триггер: Ячейка содержит символ '/' и хотя бы одну цифру.
    """
    text = clean_text(cell_text)
    if not text:
        return False
    # Содержит / и цифру
    return '/' in text and bool(re.search(r'\d', text))


def is_day_name(cell_text: str) -> Optional[str]:
    """
    Check if text is a day name and return normalized day name.
    Returns None if not a day.
    """
    text = clean_text(cell_text).lower()
    if not text:
        return None

    day_names = {
        'понедельник': 'Понедельник',
        'вторник': 'Вторник',
        'среда': 'Среда',
        'четверг': 'Четверг',
        'пятница': 'Пятница',
        'суббота': 'Суббота',
        'воскресенье': 'Воскресенье',
        'пн': 'Понедельник',
        'вт': 'Вторник',
        'ср': 'Среда',
        'чт': 'Четверг',
        'пт': 'Пятница',
        'сб': 'Суббота',
        'вс': 'Воскресенье',
    }

    for rus, normalized in day_names.items():
        if rus in text:
            return normalized

    return None


def is_pair_number(cell_text: str) -> Optional[int]:
    """
    Check if cell contains a pair number (0-8).
    Триггер: Строка, где в столбце A находится цифра.
    """
    text = clean_text(cell_text)
    if text.isdigit():
        num = int(text)
        if 0 <= num <= 8:
            return num
    return None


def get_cell_value_with_merged(sheet, cell: Cell) -> str:
    """
    Получить значение ячейки с учетом объединенных ячеек.
    Если ячейка входит в объединенный диапазон, возвращает значение главной ячейки.
    """
    coordinate = cell.coordinate

    # Проверяем, входит ли ячейка в объединенный диапазон
    for merged_range in sheet.merged_cells.ranges:
        if coordinate in merged_range:
            # Берем значение из главной ячейки (верхняя-левая)
            main_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            return clean_text(main_cell.value)

    return clean_text(cell.value)


def extract_lesson_data(
    sheet,
    row_idx: int,
    pair_num: int
) -> Optional[Dict[str, Any]]:
    """
    Извлечь данные пары из строки.

    Args:
        sheet: Лист Excel
        row_idx: Индекс строки (1-based)
        pair_num: Номер пары

    Returns:
        Словарь с данными пары или None если предмет пустой
    """
    row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]

    # Столбец A - pair_num (уже получили)
    # Столбец B - lesson
    # Столбец F - teacher
    # Столбец I - classroom

    lesson = get_cell_value_with_merged(sheet, row[1]) if len(row) > 1 else ""
    teacher = get_cell_value_with_merged(sheet, row[5]) if len(row) > 5 else ""
    classroom = get_cell_value_with_merged(sheet, row[8]) if len(row) > 8 else ""

    # Пустые пары игнорируем
    if not lesson:
        return None

    # Обработка переносов строк - заменяем на ", "
    teacher = teacher.replace('\n', ', ').replace('\r', '')
    classroom = classroom.replace('\n', ', ').replace('\r', '')

    return {
        "pair_num": pair_num,
        "lesson": lesson,
        "teacher": teacher,
        "classroom": classroom
    }


def parse_pair_rows(
    sheet,
    row_idx: int,
    current_pair: int
) -> Tuple[List[Dict[str, Any]], int]:
    """
    Обработка строк пары с определением типа недели.

    Случай А (Две строки): Если следующая строка НЕ содержит цифру в столбце A,
    но содержит предмет в столбце B, это означает разделение по неделям.
        - Верхняя строка -> type: "Четная"
        - Нижняя строка -> type: "Нечетная"
        - Если обе строки идентичны -> type: "Еженедельно" (одна запись)

    Случай Б (Одна строка): Если следующая строка пустая или содержит номер пары,
    то текущая пара type: "Еженедельно".

    Returns:
        (list of lessons, number of rows to skip)
    """
    lessons = []

    # Получаем текущую строку
    current_row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]
    current_lesson = get_cell_value_with_merged(sheet, current_row[1]) if len(current_row) > 1 else ""
    current_teacher = get_cell_value_with_merged(sheet, current_row[5]) if len(current_row) > 5 else ""
    current_classroom = get_cell_value_with_merged(sheet, current_row[8]) if len(current_row) > 8 else ""

    # Проверяем следующую строку
    next_row_idx = row_idx + 1
    has_next_row = next_row_idx <= sheet.max_row

    if has_next_row:
        next_row = list(sheet.iter_rows(min_row=next_row_idx, max_row=next_row_idx))[0]
        next_col_a = clean_text(next_row[0].value) if len(next_row) > 0 else ""
        next_lesson = get_cell_value_with_merged(sheet, next_row[1]) if len(next_row) > 1 else ""
        next_teacher = get_cell_value_with_merged(sheet, next_row[5]) if len(next_row) > 5 else ""
        next_classroom = get_cell_value_with_merged(sheet, next_row[8]) if len(next_row) > 8 else ""

        next_is_pair_num = is_pair_number(next_col_a) is not None
        next_has_lesson = bool(next_lesson)

        # Случай А: следующая строка без номера пары, но с предметом
        # Это разделение по неделям
        if not next_is_pair_num and next_has_lesson and current_lesson:
            # Проверяем, одинаковые ли строки (тогда это еженедельно)
            is_identical = (
                current_lesson == next_lesson and
                current_teacher == next_teacher and
                current_classroom == next_classroom
            )

            if is_identical:
                # Одинаковые строки - еженедельная пара
                lesson_data = extract_lesson_data(sheet, row_idx, current_pair)
                if lesson_data:
                    lesson_data["type"] = "Еженедельно"
                    lessons.append(lesson_data)
                return lessons, 1  # Пропускаем следующую строку
            else:
                # Разные строки - разделение по неделям
                # Верхняя строка - Четная
                lesson_data = extract_lesson_data(sheet, row_idx, current_pair)
                if lesson_data:
                    lesson_data["type"] = "Четная"
                    lessons.append(lesson_data)

                # Нижняя строка - Нечетная
                lesson_data_odd = extract_lesson_data(sheet, next_row_idx, current_pair)
                if lesson_data_odd:
                    lesson_data_odd["type"] = "Нечетная"
                    lessons.append(lesson_data_odd)

                return lessons, 1  # Пропускаем следующую строку

        # Если верхняя строка пустая, а нижняя с предметом
        # Это случай когда четная неделя - окно, нечетная - занятие
        if not next_is_pair_num and next_has_lesson and not current_lesson:
            lesson_data = extract_lesson_data(sheet, next_row_idx, current_pair)
            if lesson_data:
                lesson_data["type"] = "Нечетная"
                lessons.append(lesson_data)
            return lessons, 1

    # Случай Б: Еженедельная пара
    lesson_data = extract_lesson_data(sheet, row_idx, current_pair)
    if lesson_data:
        lesson_data["type"] = "Еженедельно"
        lessons.append(lesson_data)

    return lessons, 0


def parse_sheet(sheet) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Parse a single sheet using FSM approach.

    Состояния:
    - "Ищу группу" -> "Ищу день" -> "Ищу пару"

    Returns:
        Словарь: {group_name: {day_name: [lessons]}}
    """
    schedules: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}
    current_group: Optional[str] = None
    current_day: Optional[str] = None

    print(f"  Parsing sheet: {sheet.title} ({sheet.max_row} rows)")

    row_idx = 1
    while row_idx <= sheet.max_row:
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        col_a = clean_text(row[0].value) if len(row) > 0 else ""

        # Поиск Группы
        if is_group_header(col_a):
            current_group = col_a.strip()
            if current_group not in schedules:
                schedules[current_group] = {}
            print(f"    Found group: {current_group}")
            row_idx += 1
            continue

        # Поиск Дня недели
        day = is_day_name(col_a)
        if day and current_group:
            current_day = day
            if current_day not in schedules[current_group]:
                schedules[current_group][current_day] = []
            print(f"      Found day: {current_day}")
            row_idx += 1
            continue

        # Поиск Пары
        pair_num = is_pair_number(col_a)
        if pair_num is not None and current_group and current_day:
            lessons, skip_rows = parse_pair_rows(sheet, row_idx, pair_num)
            if lessons:
                schedules[current_group][current_day].extend(lessons)
            row_idx += 1 + skip_rows
            continue

        row_idx += 1

    return schedules


def parse_excel_from_bytes(excel_bytes: bytes) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Parse Excel file from bytes.

    Args:
        excel_bytes: Содержимое Excel файла как bytes

    Returns:
        Словарь расписаний для всех групп
    """
    workbook = openpyxl.load_workbook(
        filename=BytesIO(excel_bytes),
        data_only=True,
        read_only=False
    )

    all_schedules: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_schedules = parse_sheet(sheet)

        # Объединяем расписания
        for group, schedule in sheet_schedules.items():
            if group in all_schedules:
                print(f"    Warning: Group {group} appears in multiple sheets, merging...")
                for day, lessons in schedule.items():
                    if day not in all_schedules[group]:
                        all_schedules[group][day] = []
                    all_schedules[group][day].extend(lessons)
            else:
                all_schedules[group] = schedule

    workbook.close()
    return all_schedules


def parse_excel_file(file_path: str) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Parse Excel file from local path.

    Args:
        file_path: Путь к Excel файлу

    Returns:
        Словарь расписаний для всех групп
    """
    with open(file_path, 'rb') as f:
        excel_bytes = f.read()
    return parse_excel_from_bytes(excel_bytes)


def parse_excel_from_url(url: str) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """
    Parse Excel file from URL (GitHub raw file).

    Args:
        url: URL для скачивания Excel файла

    Returns:
        Словарь расписаний для всех групп
    """
    try:
        import httpx
    except ImportError:
        raise ImportError("httpx not installed. Run: pip install httpx")

    print(f"Downloading from: {url}")
    response = httpx.get(url, timeout=30)
    response.raise_for_status()
    print(f"Downloaded: {len(response.content)} bytes")

    return parse_excel_from_bytes(response.content)


def convert_to_json(
    schedules: Dict[str, Dict[str, List[Dict[str, Any]]]],
    output_path: str
) -> bool:
    """
    Сохранить расписание в JSON файл.

    Args:
        schedules: Словарь расписаний
        output_path: Путь для сохранения

    Returns:
        True если успешно, False иначе
    """
    print(f"\nSaving to {output_path}...")
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(schedules, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving JSON: {e}")
        return False

    # Статистика
    total_groups = len(schedules)
    total_lessons = sum(
        len(lesson)
        for group_schedule in schedules.values()
        for lesson in group_schedule.values()
    )

    print(f"\n✅ Conversion complete!")
    print(f"   Groups: {total_groups}")
    print(f"   Total lessons: {total_lessons}")
    print(f"   Output: {output_path}")

    return True


def main():
    """Main entry point for CLI usage."""
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel_parser.py <excel_file_or_url> [output.json]")
        print("\nExamples:")
        print("  Local file:")
        print("    python excel_parser.py oit_2sem.xlsx")
        print("  From GitHub URL:")
        print("    python excel_parser.py https://raw.githubusercontent.com/user/repo/main/oit_2sem.xlsx")
        print("\nOutput format compatible with ЯГК Schedule system.")
        sys.exit(1)

    source = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'schedule.json'

    try:
        if source.startswith('http://') or source.startswith('https://'):
            schedules = parse_excel_from_url(source)
        else:
            from pathlib import Path
            if not Path(source).exists():
                print(f"Error: File not found: {source}")
                sys.exit(1)
            schedules = parse_excel_file(source)

        if not schedules:
            print("Error: No schedule data parsed")
            sys.exit(1)

        success = convert_to_json(schedules, output_path)
        sys.exit(0 if success else 1)

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
