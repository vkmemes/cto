#!/usr/bin/env python3
import json
import sys
import re
import os
from pathlib import Path
from typing import Dict, List, Optional, Any

try:
    import openpyxl
    from openpyxl.cell import Cell
except ImportError:
    print("❌ Ошибка: Библиотека openpyxl не установлена.")
    print("👉 Выполните: pip install openpyxl")
    sys.exit(1)

# --- КОНФИГУРАЦИЯ КОЛОНОК (A=1, B=2, F=6, I=9) ---
COL_PAIR = 1    # Номер пары
COL_SUBJ = 2    # Предмет
COL_TEACH = 6   # Преподаватель
COL_ROOM = 9    # Аудитория

# --- ОТЛАДКА ---
# Включи это, чтобы видеть лог по каждой строке в консоли
DEBUG_MODE = True 

def clean_text(text) -> str:
    """Убирает лишние пробелы и переносы"""
    if text is None: return ""
    text = str(text).strip()
    return re.sub(r'\s+', ' ', text)

def get_pair_number(value) -> Optional[int]:
    """Извлекает номер пары (0, 1, 2...)"""
    if value is None: return None
    s = str(value).strip()
    if s.isdigit(): return int(s)
    match = re.match(r'^(\d+)', s) # Если там "1 пара"
    if match: return int(match.group(1))
    return None

def is_group_header(text: str) -> bool:
    """Это заголовок группы?"""
    if not text: return False
    # Исключаем дни недели
    if any(d in text.lower() for d in ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота']):
        return False
    # Ищем наличие букв и цифр (ИС1-21)
    return bool(re.search(r'[А-ЯA-Z].*\d', text)) and len(text) < 50

def get_day_id(text: str) -> Optional[str]:
    """Возвращает нормализованное название дня"""
    if not text: return None
    t = text.lower()
    days = {
        'понедельник': 'Понедельник', 'вторник': 'Вторник', 'среда': 'Среда',
        'четверг': 'Четверг', 'пятница': 'Пятница', 'суббота': 'Суббота', 'воскресенье': 'Воскресенье'
    }
    for ru, normal in days.items():
        if ru in t: return normal
    return None

def get_merged_range(sheet, row, col):
    """Возвращает диапазон объединения, если ячейка входит в него, иначе None"""
    cell = sheet.cell(row=row, column=col)
    for merged in sheet.merged_cells.ranges:
        if cell.coordinate in merged:
            return merged
    return None

def get_val(sheet, row, col):
    """
    Умное получение значения. 
    Если ячейка часть объединения, берет значение из левой верхней.
    НО! Нам иногда нужно знать реальное значение конкретной ячейки (даже если она скрыта объединением).
    Для openpyxl value есть только у левой верхней.
    """
    cell = sheet.cell(row=row, column=col)
    # Если это просто ячейка
    if cell.value is not None:
        return clean_text(cell.value)
    
    # Если значение None, проверяем, не перекрыто ли оно объединением
    merged = get_merged_range(sheet, row, col)
    if merged:
        # Берем значение из главной ячейки диапазона
        val = sheet.cell(row=merged.min_row, column=merged.min_col).value
        return clean_text(val)
    
    return ""

def parse_sheet(sheet) -> Dict[str, Dict[str, List[Dict]]]:
    print(f"  📄 Чтение листа: {sheet.title}...")
    
    sheet_schedule = {} # { "Группа": { "Понедельник": [] } }
    current_groups = [] # Теперь это список из одной строки, например ["СА1-11/СА1-12"]
    current_day = None
    
    row_idx = 1
    max_row = sheet.max_row
    
    while row_idx <= max_row:
        # Читаем "сырое" значение колонки A (без учета объединений для начала)
        raw_cell_a = sheet.cell(row=row_idx, column=COL_PAIR)
        val_a = clean_text(raw_cell_a.value)
        
        # --- 1. Поиск Группы ---
        # Если это объединенная ячейка с текстом группы
        # Берем значение через get_val, чтобы достать текст из merge
        real_text_a = get_val(sheet, row_idx, COL_PAIR)
        
        if is_group_header(real_text_a):
            # ВАЖНО: Берем название как есть, не сплитим по слэшам
            # Убираем лишние переносы
            group_name = real_text_a.replace('\n', ' ').strip()
            current_groups = [group_name]
            
            # Инициализация
            if group_name not in sheet_schedule:
                sheet_schedule[group_name] = {}
            
            if DEBUG_MODE: print(f"📍 [Row {row_idx}] Найдена группа: {group_name}")
            current_day = None
            row_idx += 1
            continue

        # --- 2. Поиск Дня ---
        day_name = get_day_id(real_text_a)
        if day_name:
            current_day = day_name
            for g in current_groups:
                if current_day not in sheet_schedule[g]:
                    sheet_schedule[g][current_day] = []
            if DEBUG_MODE: print(f"🗓 [Row {row_idx}] День: {current_day}")
            row_idx += 1
            continue

        # --- 3. Поиск Пары ---
        # Проверяем, есть ли номер пары в ячейке.
        # ВАЖНО: Если ячейка объединена вертикально, то значение есть только в верхней строке.
        # В нижней строке value будет None.
        
        # Получаем диапазон объединения для номера пары
        merged_range_a = get_merged_range(sheet, row_idx, COL_PAIR)
        
        # Если мы попали на "нижнюю" часть объединенной ячейки номера (которая пустая),
        # мы её пропускаем, так как обработали на шаге "верхней".
        # Но наша логика ниже (jump +2) должна это предотвращать.
        
        # Определяем номер пары. Берем value именно этой ячейки.
        pair_num = get_pair_number(raw_cell_a.value)
        
        if pair_num is not None and current_groups and current_day:
            
            # Определяем, сколько строк занимает эта пара
            rows_span = 1
            if merged_range_a:
                # Считаем высоту объединения
                rows_span = merged_range_a.max_row - merged_range_a.min_row + 1
            
            # Данные верхней строки (Четная / Еженедельная)
            subj_top = get_val(sheet, row_idx, COL_SUBJ)
            teach_top = get_val(sheet, row_idx, COL_TEACH)
            room_top = get_val(sheet, row_idx, COL_ROOM)
            
            if rows_span >= 2:
                # --- СЦЕНАРИЙ: Пара занимает 2 строки ---
                # Проверяем, объединен ли ПРЕДМЕТ (Col B)
                merged_range_b = get_merged_range(sheet, row_idx, COL_SUBJ)
                
                is_subject_merged = False
                if merged_range_b:
                    # Если предмет объединен на те же (или больше) строк, что и номер
                    if (merged_range_b.max_row - merged_range_b.min_row + 1) >= 2:
                        is_subject_merged = True

                if is_subject_merged:
                    # >>> ЕЖЕНЕДЕЛЬНО (И номер, и предмет объединены)
                    if DEBUG_MODE: print(f"   [Row {row_idx}] Пара {pair_num}: ЕЖЕНЕДЕЛЬНО (Merged)")
                    if subj_top:
                        lesson = {"pair_num": pair_num, "type": "Еженедельно", "lesson": subj_top, "teacher": teach_top, "classroom": room_top}
                        for g in current_groups: sheet_schedule[g][current_day].append(lesson)
                
                else:
                    # >>> РАЗДЕЛЕНИЕ (Номер объединен, а предметы разные)
                    # Верхняя строка (row_idx) = Четная
                    # Нижняя строка (row_idx + 1) = Нечетная
                    
                    # Данные нижней строки
                    subj_bot = get_val(sheet, row_idx + 1, COL_SUBJ)
                    teach_bot = get_val(sheet, row_idx + 1, COL_TEACH)
                    room_bot = get_val(sheet, row_idx + 1, COL_ROOM)

                    if DEBUG_MODE: print(f"   [Row {row_idx}] Пара {pair_num}: РАЗДЕЛЕНИЕ")
                    
                    # Добавляем Четную (Верх)
                    if subj_top:
                        if DEBUG_MODE: print(f"     -> Четная: {subj_top}")
                        l_even = {"pair_num": pair_num, "type": "Четная", "lesson": subj_top, "teacher": teach_top, "classroom": room_top}
                        for g in current_groups: sheet_schedule[g][current_day].append(l_even)
                    
                    # Добавляем Нечетную (Низ)
                    if subj_bot:
                        if DEBUG_MODE: print(f"     -> Нечетная: {subj_bot}")
                        l_odd = {"pair_num": pair_num, "type": "Нечетная", "lesson": subj_bot, "teacher": teach_bot, "classroom": room_bot}
                        for g in current_groups: sheet_schedule[g][current_day].append(l_odd)

                # Пропускаем столько строк, сколько занимал номер пары
                row_idx += rows_span
            
            else:
                # --- СЦЕНАРИЙ: Пара занимает 1 строку ---
                # Это обычная Еженедельная пара
                if DEBUG_MODE: print(f"   [Row {row_idx}] Пара {pair_num}: ЕЖЕНЕДЕЛЬНО (Single)")
                if subj_top:
                    lesson = {"pair_num": pair_num, "type": "Еженедельно", "lesson": subj_top, "teacher": teach_top, "classroom": room_top}
                    for g in current_groups: sheet_schedule[g][current_day].append(lesson)
                
                row_idx += 1
            
            continue

        # Пустая строка или мусор
        row_idx += 1
        
    return sheet_schedule

def main():
    # ... (стандартный код поиска файлов и сохранения) ...
    script_dir = Path(__file__).parent
    files = sorted(script_dir.glob('*.xlsx'))
    
    if not files:
        print("❌ .xlsx файлы не найдены!")
        sys.exit(1)
        
    final_schedule = {}
    
    for excel_file in files:
        if excel_file.name.startswith("~"): continue
        print(f"\n📂 Обработка файла: {excel_file.name}")
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet_data = parse_sheet(wb[sheet_name])
                # Простое слияние словарей
                final_schedule.update(sheet_data)
            wb.close()
        except Exception as e:
            print(f"❌ Ошибка: {e}")

    output_path = script_dir / 'schedule.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(final_schedule, f, ensure_ascii=False, indent=4)
        
    print(f"\n💾 Сохранено в: {output_path}")

if __name__ == "__main__":
    main()