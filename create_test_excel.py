#!/usr/bin/env python3
"""
Create test Excel file with 32 groups for testing excel_to_schedule.py
"""

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Расписание"

# Groups (32 groups as requested)
groups = [
    "ИС1-11", "ИС1-12", "ИС1-13", "ИС1-14",
    "ИС2-11", "ИС2-12", "ИС2-13", "ИС2-14",
    "ПО1-11", "ПО1-12", "ПО1-13", "ПО1-14",
    "ПО2-11", "ПО2-12", "ПО2-13", "ПО2-14",
    "СА1-11", "СА1-12", "СА1-13", "СА1-14",
    "СА2-11", "СА2-12", "СА2-13", "СА2-14",
    "ТО1-11", "ТО1-12", "ТО1-13", "ТО1-14",
    "ТО2-11", "ТО2-12", "ТО2-13", "ТО2-14"
]

# Time slots
time_slots = [
    "08:30-10:00",
    "10:10-11:40",
    "12:00-13:30",
    "13:40-15:10",
    "15:20-16:50"
]

# Sample subjects
subjects = [
    "Математика",
    "Информатика",
    "Русский язык",
    "История",
    "Английский язык",
    "Физика",
    "Химия",
    "Физкультура",
    "Программирование",
    "Базы данных",
    "Web-технологии",
    "Компьютерные сети",
    "Операционные системы"
]

teachers = [
    "Иванов И.И.",
    "Петров П.П.",
    "Сидорова С.С.",
    "Кузнецов К.К.",
    "Смирнова А.А.",
    "Новиков Н.Н.",
    "Федорова Ф.Ф."
]

rooms = ["201", "202", "203", "305", "306", "307", "Спортзал", "Актовый зал"]

days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
weeks = ["Числитель", "Знаменатель"]

# Header styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
day_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
day_font = Font(bold=True)

# Create header row with groups
ws['A1'] = "Время"
ws['A1'].font = header_font
ws['A1'].fill = header_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

for col_idx, group in enumerate(groups, start=2):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = group
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

current_row = 2

# Generate schedule for each week and day
for week in weeks:
    for day in days:
        # Day header
        ws.cell(row=current_row, column=1).value = f"{day} ({week})"
        ws.cell(row=current_row, column=1).font = day_font
        ws.cell(row=current_row, column=1).fill = day_fill
        current_row += 1
        
        # Time slots for this day
        for time_idx, time_slot in enumerate(time_slots):
            # Time column
            ws.cell(row=current_row, column=1).value = time_slot
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Lessons for each group
            for col_idx, group in enumerate(groups, start=2):
                # Generate some lessons (not all slots filled)
                import random
                if random.random() < 0.7:  # 70% chance of having a lesson
                    subject = random.choice(subjects)
                    teacher = random.choice(teachers)
                    room = random.choice(rooms)
                    
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = f"{subject}\n{teacher}\n{room}"
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            current_row += 1
        
        # Empty row after each day
        current_row += 1

# Adjust column widths
ws.column_dimensions['A'].width = 15
for col_idx in range(2, len(groups) + 2):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

# Adjust row heights
for row in ws.iter_rows(min_row=2):
    ws.row_dimensions[row[0].row].height = 60

# Save
filename = "test_schedule_32groups.xlsx"
wb.save(filename)
print(f"✅ Created test Excel file: {filename}")
print(f"   Groups: {len(groups)}")
print(f"   Days: {len(days)}")
print(f"   Weeks: {len(weeks)}")
print(f"   Time slots per day: {len(time_slots)}")
print(f"\nTest conversion with:")
print(f"   python excel_to_schedule.py {filename}")
